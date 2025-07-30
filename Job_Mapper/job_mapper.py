import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# === PATHS ===
input_folder = os.getcwd()
output_folder = os.path.join(input_folder, "1")
mappings_folder = os.path.join(input_folder, "Mappings")
name_map_file = os.path.join(input_folder, "name_map.xlsx")
mapping_issues_file = os.path.join(output_folder, "phase_mapping_issues.xlsx")

# === CONFIG ===
exclude_phase_codes = {"A0000.01021000", "A0000.01022000", "A0000.00000000", "9999"}
required_columns = ["New Phase", "New Phase Description"]
type3_new_phase = "B1000.01700000"
type3_new_phase_description = "Execution/Closeout Requirement"

regular_cost_type_mapping = {
    "B": "Materials", "C": "Materials", "E": "Equipment", "I": "Equipment", "IE": "Equipment",
    "IL": "Labour", "L": "Labour", "M": "Materials", "S": "Subcontracting", "T": "Subcontracting",
    "R": "Materials", "P": "Materials", "IND": "Materials"
}

type3_cost_type_mapping = {
    "1": "Labor",
    "2": "Materials",
    "3": "Equipment",
    "4": "Subcontracting",
    "5": "Sundry"
}

os.makedirs(output_folder, exist_ok=True)

# === LOAD NAME MAP ===
df_name_map = pd.read_excel(name_map_file, dtype=str)
name_mapping_dict = dict(zip(df_name_map["Old Name"].str.strip(), df_name_map["New Name"].str.strip()))

# === TRACKING MISSING MAPPINGS ===
missing_mapping_files = set()
incomplete_mapping_values = set()

def map_job_number(job_number):
    return name_mapping_dict.get(job_number.strip(), job_number.strip())

def load_mapping_table(job_number, is_type3=False):
    if is_type3:
        return {
            "New Phase": type3_new_phase,
            "New Phase Description": type3_new_phase_description
        }

    mapping_path = os.path.join(mappings_folder, f"m{job_number}.xlsx")
    if not os.path.exists(mapping_path):
        return None

    df_map = pd.read_excel(mapping_path, dtype=str)
    df_map.columns = df_map.columns.str.strip()
    if "PHASE" not in df_map.columns:
        return {}

    df_map["PHASE"] = df_map["PHASE"].astype(str).str.strip()
    if df_map["PHASE"].duplicated().any():
        df_map = df_map.groupby("PHASE", as_index=False).first()

    mapping_dict = df_map.set_index("PHASE")[required_columns].to_dict(orient="index")
    return mapping_dict

def map_phase_values(phase_code, mapping_dict, job_number):
    code = str(phase_code).strip()
    if code in mapping_dict:
        return mapping_dict[code]
    else:
        incomplete_mapping_values.add((job_number, code))
        return {col: None for col in required_columns}

def process_file(file_path):
    filename = os.path.basename(file_path)
    is_type3 = "_type3" in filename.lower()

    print(f"\n Processing: {filename} | Type 3: {is_type3}")
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()

    if "Job_Number" not in df.columns:
        print("'Job_Number' column is required. Skipping.")
        return

    if not is_type3 and "Phase_Code" not in df.columns:
        print("'Phase_Code' is missing in a regular file. Skipping.")
        return

    if not is_type3:
        df["Phase_Code"] = df["Phase_Code"].astype(str).str.strip()
        df = df[~df["Phase_Code"].isin(exclude_phase_codes)]
        df = df[~df["Phase_Code"].str.startswith("WIP")]
        if "Cost_Type" in df.columns:
            df = df[df["Cost_Type"] != "$"]

    # Format Date for both types
    date_column = "Date" if is_type3 else "LN Tran Date"
    if date_column in df.columns:
        df[date_column] = pd.to_datetime(df[date_column], errors='coerce')

    # Apply cost type mapping
    if is_type3 and "New Cost Type" in df.columns:
        df["New Cost Type"] = df["New Cost Type"].map(type3_cost_type_mapping).fillna(df["New Cost Type"])
    elif not is_type3 and "Cost_Type" in df.columns:
        df["New Cost Type"] = df["Cost_Type"].apply(
            lambda x: regular_cost_type_mapping.get(str(x).strip().upper(), "Sundry Costs")
        )

    # Apply New Job Number mapping
    df["New Job Number"] = df["Job_Number"].apply(map_job_number)

    # Load mappings
    job_numbers = df["Job_Number"].unique()
    mapping_dicts = {}

    for job in job_numbers:
        mapping = load_mapping_table(job, is_type3)
        df_job = df[df["Job_Number"] == job]

        if mapping is None:
            for code in df_job.get("Phase_Code", []):
                missing_mapping_files.add((job, code))
            mapping_dicts[job] = {}
        else:
            mapping_dicts[job] = mapping

    def apply_mapping(row):
        job = row["Job_Number"]
        if is_type3:
            return {
                "New Phase": type3_new_phase,
                "New Phase Description": type3_new_phase_description
            }
        return map_phase_values(row["Phase_Code"], mapping_dicts.get(job, {}), job)

    df_mapped = pd.DataFrame(df.apply(apply_mapping, axis=1).tolist())
    df.reset_index(drop=True, inplace=True)
    df_mapped.reset_index(drop=True, inplace=True)

    df_final = pd.concat([df, df_mapped], axis=1)
    output_path = os.path.join(output_folder, filename.replace(".xlsx", "_updated.xlsx"))
    df_final.to_excel(output_path, index=False)

    # Apply formatting
    if date_column in df.columns:
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_idx = headers.index(date_column) + 1
            date_style = NamedStyle(name="datetime", number_format="MM/DD/YYYY HH:MM:SS")
            if date_style.name not in wb.named_styles:
                wb.add_named_style(date_style)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.style = date_style
            wb.save(output_path)
            print(f"Date formatted in '{date_column}'")
        except Exception as e:
            print(f"Date formatting error: {e}")

    print(f"Saved: {output_path}")

# === MAIN LOOP ===
for fname in os.listdir(input_folder):
    if fname.endswith(".xlsx") and not fname.startswith("m") and fname != os.path.basename(name_map_file):
        process_file(os.path.join(input_folder, fname))

# === FINAL REPORT ===
df_missing_file = pd.DataFrame(list(missing_mapping_files), columns=["Job_Number", "Phase_Code"])
df_incomplete_map = pd.DataFrame(list(incomplete_mapping_values), columns=["Job_Number", "Phase_Code"])

df_missing_file.sort_values(by=["Job_Number", "Phase_Code"], inplace=True)
df_incomplete_map.sort_values(by=["Job_Number", "Phase_Code"], inplace=True)

if not df_missing_file.empty or not df_incomplete_map.empty:
    with pd.ExcelWriter(mapping_issues_file, engine="openpyxl") as writer:
        if not df_missing_file.empty:
            df_missing_file.to_excel(writer, index=False, sheet_name="No Mapping file")
        if not df_incomplete_map.empty:
            df_incomplete_map.to_excel(writer, index=False, sheet_name="Incomplete mappings")
    print(f"\n Mapping issues saved to: {mapping_issues_file}")

print("\n All files processed.")
